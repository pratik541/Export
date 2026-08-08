"""Builds and writes the Fantasy File (the source HTML tool's "Open Stock"
export, renamed per user decision) -- a wide layout with fixed C1/C2 metal+
labor slots and 20 further stone slots (C3-C22). Column layout and header
text are dictated by the receiving system and are NOT configurable.

Cell coloring matches a real reference "Open Stock" file's actual cell
formatting exactly for the header row's core/C1/C2/C3-C7 columns (see
_STONE_SLOT_COLORS) -- not the source HTML tool's dead OS_COLORS map (never
applied anywhere in that tool's own code) or this port's earlier
placeholder highlight scheme (chosen before any real file had been
inspected, since reproducing the source tool's actual 238-entry hardcoded
style-index output was judged not worth the effort). Once real reference
files were available, both turned out not to match production at all, so
the colors here were replaced with the real, verified values instead.

Three more deliberate departures from the source tool's literal code, all
made after comparing this port's output against real reference "Open
Stock" files (Data/JNE016, JNE008, JNE009, JNE017 Open Stock.xls,
local-only, real business data), which the source tool's code did not
reproduce:
- LotName: the source tool computes settingCert || igiCert || sn (a
  certificate-number fallback chain). The real reference file shows the
  style number (sn) on every row checked, even when a certificate was
  available -- so LotName here is always sn (user decision).
- C1:Weight: the source tool rounds to 2 decimal places
  (Math.round(w*100)/100). The real reference file shows the raw,
  unrounded packing-list value -- so no rounding is applied here (user
  decision).
- Material/Metal codes: see export_tool/materials.py and config.py's
  FANTASY_MATERIAL_MAP -- the source file's default material table
  disagreed with real production output on most KT codes across 4 real
  shipments (e.g. "14KTR" -> "14KT RG" per the source file's hardcoded
  default vs. "14KR" in every real reference checked)."""
from io import BytesIO

import openpyxl
from openpyxl.styles import PatternFill

from export_tool import config, materials, stones
from export_tool._util import is_blank, master_style, safe_num

MAX_SLOTS = 22

_CORE_COLUMNS = [
    "PO Doc ID", "PO Doc Line", "Item Name", "LotName", "Metal", "Order #",
    "Qty", "Weight", " Total H.Cost", "ItemTypeID",
    "C1:Item Name", "C1:LotName", "C1:Weight", "C1: Total H.Cost",
    "C2:Item Name", "C2:LotName", "C2: Qty", "C2: Total H.Cost",
]


def _stone_columns(slot: int) -> list:
    position_col = f"C{slot}: Stone Position" if slot == 3 else f"C{slot}:Stone Position"
    return [
        f"C{slot}:Item Name", f"C{slot}:LotName", f"C{slot}: Qty", f"C{slot}:Weight",
        f"C{slot}:Shape", f"C{slot}:Color", f"C{slot}:Clarity", f"C{slot}: Lab",
        f"C{slot}:Cert#", f"C{slot}: Total H.Cost", position_col,
    ]


FANTASY_COLUMNS = list(_CORE_COLUMNS)
for _slot in range(3, MAX_SLOTS + 1):
    FANTASY_COLUMNS.extend(_stone_columns(_slot))

_NUMERIC_COLUMNS = {
    "Qty", "Weight", " Total H.Cost", "ItemTypeID",
    "C1:Weight", "C1: Total H.Cost", "C2: Qty", "C2: Total H.Cost",
}
for _slot in range(3, MAX_SLOTS + 1):
    _NUMERIC_COLUMNS.update({f"C{_slot}: Qty", f"C{_slot}:Weight", f"C{_slot}: Total H.Cost"})

_AVAILABLE_SLOTS = MAX_SLOTS - 3 + 1


def build_rows(items: list, jobsheet_index: dict) -> tuple:
    rows = []
    warnings = []
    unmatched_jobsheet_items = []
    unmapped_kt_codes = {}  # kt code -> list of style numbers that used it
    no_side_categories = {c.strip().upper() for c in config.NO_SIDE_DIAMOND_CATEGORIES}

    for item in items:
        style_no = item["sn"]
        jobsheet_row = (
            jobsheet_index.get(style_no)
            or jobsheet_index.get(master_style(style_no))
            or {}
        )
        if not jobsheet_row:
            unmatched_jobsheet_items.append(style_no)
        fantasy = materials.resolve_fantasy_material(item["kt"])
        if not fantasy.matched:
            unmapped_kt_codes.setdefault(item["kt"], []).append(style_no)
        design_no = _jobsheet_value(jobsheet_row, config.JOBSHEET_COLUMNS["design_no"])
        parent_style = _jobsheet_value(jobsheet_row, config.JOBSHEET_COLUMNS["parent_style"])

        groups = stones.aggregate(item["stones"])
        if (item["cat"] or "").strip().upper() in no_side_categories:
            groups = [g for g in groups if g["position"] == "center"]

        row = {column: None for column in FANTASY_COLUMNS}
        row["Item Name"] = materials.build_item_name(parent_style, master_style(style_no), fantasy.suffix)
        row["LotName"] = style_no
        row["Metal"] = fantasy.metal
        row["Order #"] = design_no
        row["Qty"] = safe_num(item["qty"])
        row["Weight"] = safe_num(item["gw"])
        row[" Total H.Cost"] = 0
        row["ItemTypeID"] = config.ITEM_TYPE_ID
        row["C1:Item Name"] = fantasy.c1
        row["C1:LotName"] = fantasy.c1
        row["C1:Weight"] = safe_num(item["tmw"])
        row["C1: Total H.Cost"] = safe_num(item["mv"])
        row["C2:Item Name"] = "Labor"
        row["C2:LotName"] = "Labor"
        row["C2: Qty"] = safe_num(item["qty"])
        row["C2: Total H.Cost"] = safe_num(item["making"])

        if len(groups) > _AVAILABLE_SLOTS:
            warnings.append(
                f'Item "{style_no}": {len(groups) - _AVAILABLE_SLOTS} stone group(s) '
                f"truncated to fit C3-C{MAX_SLOTS}."
            )

        for slot_index, group in enumerate(groups[:_AVAILABLE_SLOTS]):
            slot = slot_index + 3
            is_center = group["position"] == "center"
            lot = group["cert"] if (is_center and not is_blank(group["cert"])) else "Diamonds"
            row[f"C{slot}:Item Name"] = "Diamonds"
            row[f"C{slot}:LotName"] = lot
            row[f"C{slot}: Qty"] = safe_num(group["pcs"])
            row[f"C{slot}:Weight"] = safe_num(group["cts"])
            row[f"C{slot}:Shape"] = group["shape"]
            row[f"C{slot}:Color"] = materials.normalize_os_color(group["color"])
            row[f"C{slot}:Clarity"] = group["clarity"]
            row[f"C{slot}: Lab"] = group["lab"] if (is_center and not is_blank(group["lab"])) else None
            row[f"C{slot}:Cert#"] = group["cert"] if (is_center and not is_blank(group["cert"])) else None
            row[f"C{slot}: Total H.Cost"] = safe_num(group["val"])
            position_col = f"C{slot}: Stone Position" if slot == 3 else f"C{slot}:Stone Position"
            row[position_col] = "C" if is_center else "S"

        rows.append(row)

    if unmatched_jobsheet_items:
        warnings.append(
            f"{len(unmatched_jobsheet_items)} item(s) had no matching jobsheet row "
            f"(Order #/parent style left blank): {', '.join(unmatched_jobsheet_items)}"
        )
    if unmapped_kt_codes:
        total = sum(len(styles) for styles in unmapped_kt_codes.values())
        codes = ", ".join(
            f'"{kt}" ({len(styles)}x, e.g. {styles[0]})' for kt, styles in sorted(unmapped_kt_codes.items())
        )
        warnings.append(
            f"{total} item(s) used a KT code not in FANTASY_MATERIAL_MAP (raw code used "
            f"as-is for Metal/C1): {codes} -- add to export_tool/config.py if these should map to something else."
        )

    return rows, warnings


def _jobsheet_value(jobsheet_row: dict, column_name: str):
    value = jobsheet_row.get(column_name)
    if value is None:
        return None
    text = str(value).strip()
    return text or None


_GRAY = "C0C0C0"
_LIGHT_YELLOW = "FFFF99"
_BLUE = "00CCFF"
_PEACH = "FFCC99"
_BRIGHT_YELLOW = "FFFF00"
_OLIVE = "99CC00"

# Stone-slot fill colors, extracted directly from real reference "Open
# Stock" files' actual cell formatting (Data/JNE016, JNE008, JNE009, JNE017
# Open Stock.xls, local-only, real business data -- all 4 agree exactly on
# slots C1-C7). Slots C8-C22 have no verified color: no real shipment
# checked used more than 5 distinct stone groups (slot C7), so they're left
# unfilled rather than guessed.
_STONE_SLOT_COLORS = {3: _PEACH, 4: _LIGHT_YELLOW, 5: _BRIGHT_YELLOW, 6: _OLIVE, 7: _OLIVE}


def _column_fills() -> dict:
    """1-indexed column -> fill hex color, matching a real reference file's
    actual formatting exactly (see _STONE_SLOT_COLORS) -- not the source
    HTML tool's dead, never-applied OS_COLORS map, which was checked against
    real files and does not match production output at all (e.g. its C4 is
    bright yellow FFFF00; real C4 is light yellow FFFF99, the same as C1)."""
    fills = {}
    for col in range(3, 11):
        fills[col] = _GRAY
    for col in range(11, 15):
        fills[col] = _LIGHT_YELLOW
    for col in range(15, 19):
        fills[col] = _BLUE
    for slot, color in _STONE_SLOT_COLORS.items():
        start = 19 + 11 * (slot - 3)
        for col in range(start, start + 11):
            fills[col] = color
    return fills


def write_xlsx(rows: list) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    fills = _column_fills()

    for col_index, column in enumerate(FANTASY_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=column)
        if col_index in fills:
            color = fills[col_index]
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    for row_index, row in enumerate(rows, start=2):
        for col_index, column in enumerate(FANTASY_COLUMNS, start=1):
            value = row.get(column)
            if value is None:
                continue
            if column in _NUMERIC_COLUMNS:
                value = safe_num(value)
                if value is None:
                    continue
            sheet.cell(row=row_index, column=col_index, value=value)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
