"""Regression test: runs the real JNE016 sample packing list through
export_tool.packing_list.parse_packing_list and checks it parses cleanly
against the column headers configured in export_tool/config.py.

The sample file lives in Data/ and is intentionally NOT committed to git
(real shipment data). This test skips itself when it's absent, so the suite
still passes in a fresh clone or CI -- same pattern as
tests/test_working_sheet_real_sample.py."""
from pathlib import Path

import pytest

from export_tool import fantasy_file, packing_list

DATA_DIR = Path(__file__).parent.parent / "Data"
PACKING_LIST_PATH = DATA_DIR / "JNE016 CR Packing List Export Report new.xlsx"

pytestmark = pytest.mark.skipif(
    not PACKING_LIST_PATH.exists(),
    reason="Real sample shipment file is local-only (not committed) -- skipped when absent.",
)


def test_real_packing_list_parses_with_no_missing_columns():
    items, warnings = packing_list.parse_packing_list(PACKING_LIST_PATH.read_bytes())

    assert warnings == []
    assert len(items) > 0
    assert all(item["sn"] for item in items)
    assert all(item["kt"] for item in items)


def test_real_packing_list_has_a_bracelet_item():
    # Confirms the real file exercises NO_SIDE_DIAMOND_CATEGORIES filtering
    # (applied downstream in fantasy_file.build_rows).
    items, _ = packing_list.parse_packing_list(PACKING_LIST_PATH.read_bytes())
    assert any((item["cat"] or "").upper() == "BRACELET" for item in items)


def test_real_packing_list_flows_through_the_full_pipeline_to_a_valid_workbook():
    from io import BytesIO
    import openpyxl

    items, warnings = packing_list.parse_packing_list(PACKING_LIST_PATH.read_bytes())
    assert warnings == []

    rows, build_warnings = fantasy_file.build_rows(items, {})
    assert build_warnings == []
    assert len(rows) == len(items)

    workbook_bytes = fantasy_file.write_xlsx(rows)
    workbook = openpyxl.load_workbook(BytesIO(workbook_bytes))
    sheet = workbook.active
    assert sheet.max_column == len(fantasy_file.FANTASY_COLUMNS) == 238
    assert sheet.max_row == len(rows) + 1  # +1 for the header row

    # Regression pin for the exact production-parity fix: these two KT codes
    # are both present in this real file and must resolve exactly as today's
    # live tool does (see export_tool/materials.py's _SPECIAL_CASE_TRIGGERS).
    kt_codes = {item["kt"] for item in items}
    assert "14KTR" in kt_codes
    assert "14KTY" in kt_codes

    metal_col = fantasy_file.FANTASY_COLUMNS.index("Metal") + 1
    metals_seen = {sheet.cell(row=r, column=metal_col).value for r in range(2, sheet.max_row + 1)}
    assert "14KT RG" in metals_seen   # from the 14KTR item -- must be mapped
    assert "14KTY" in metals_seen     # from the 14KTY items -- must stay unmapped
