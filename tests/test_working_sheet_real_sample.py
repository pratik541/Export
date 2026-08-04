"""End-to-end regression test: runs the real JNE016 sample shipment through
the full Packing List + Invoice -> Working Sheet pipeline and checks the
output against the real, hand-built Working Sheet, row for row.

The sample files live in Data/ and are intentionally NOT committed to git
(they contain real business/financial data -- bank account number, GST/PAN/
IEC numbers, a real customer name). This test skips itself when they aren't
present, so the suite still passes in a fresh clone or CI; it's a local-only
regression guard for whoever has the files."""
import math
from pathlib import Path

import openpyxl
import pytest

from working_sheet import builder, invoice as invoice_parser, packing_list

DATA_DIR = Path(__file__).parent.parent / "Data"
PACKING_LIST_PATH = DATA_DIR / "JNE016 CR Packing List Export Report new.xlsx"
INVOICE_PATH = DATA_DIR / "JNE016 CR Export Invoice Report New.pdf"
WORKING_SHEET_PATH = DATA_DIR / "Working Sheet JNE016-26-27.xlsx"

pytestmark = pytest.mark.skipif(
    not (PACKING_LIST_PATH.exists() and INVOICE_PATH.exists() and WORKING_SHEET_PATH.exists()),
    reason="Real sample shipment files are local-only (not committed) -- skipped when absent.",
)


def _expected_rows():
    wb = openpyxl.load_workbook(WORKING_SHEET_PATH, data_only=True)
    ws = wb.worksheets[0]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return [dict(zip(header, row)) for row in ws.iter_rows(min_row=2, values_only=True)]


def test_generated_working_sheet_matches_the_real_sample_row_for_row():
    categories = packing_list.parse_packing_list(PACKING_LIST_PATH.read_bytes())
    invoice_data = invoice_parser.parse_invoice(INVOICE_PATH.read_bytes())
    got_rows = builder.build_rows(categories, invoice_data)
    want_rows = _expected_rows()

    assert len(got_rows) == len(want_rows) == 6

    for got, want in zip(got_rows, want_rows):
        assert got["RITC"] == str(want["RITC"])
        assert got["Item Description"] == want["Item Description"]
        assert math.isclose(got["Qty."], want["Qty."], rel_tol=1e-6)
        assert math.isclose(got["Unit Price"], want["Unit Price"], rel_tol=1e-4)
        assert math.isclose(got["Standard Qty"], want["Standard Qty"], abs_tol=1e-6)
        assert got["Invoice No."] == want["Invoice No."]
        assert int(got["State Code"]) == want["State Code"]
        assert int(got["District Code"]) == want["District Code"]
        assert got["FTA Code"] == want["FTA Code"]
        assert got["IGST Payment Status"] == want["IGST Payment Status"]
        assert got["RoDTEP"] == want["RoDTEP"]


def test_cross_validation_finds_no_mismatches_on_the_real_sample():
    categories = packing_list.parse_packing_list(PACKING_LIST_PATH.read_bytes())
    invoice_data = invoice_parser.parse_invoice(INVOICE_PATH.read_bytes())

    assert builder.cross_validate(categories, invoice_data) == []
