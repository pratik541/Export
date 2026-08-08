from export_tool import fantasy_file


def _item(**overrides):
    base = {
        "sr": 1, "sn": "STYLE1", "cat": "RING", "kt": "18KT WG",
        "qty": 1, "gw": 5.0, "tmw": 4.5, "mv": 200.0, "making": 30.0,
        "cert": None, "stones": [],
    }
    base.update(overrides)
    return base


def _stone(**overrides):
    base = {
        "position": "center", "label": "LG Diamond", "shape": "RD", "color": "FG",
        "clarity": "VS", "lab": "IGI", "cert": "CERT001", "cts": 0.3, "pcs": 10, "val": 50.0,
    }
    base.update(overrides)
    return base


def test_build_rows_maps_core_fields_from_item_and_jobsheet():
    item = _item(stones=[_stone()])
    jobsheet_index = {
        "STYLE1": {
            "Order Id": "ORDER123",
            "Product Status": "STYLE1",
        }
    }

    rows, warnings = fantasy_file.build_rows([item], jobsheet_index)

    assert warnings == []
    assert len(rows) == 1
    row = rows[0]
    assert row["LotName"] == "STYLE1"
    assert row["Metal"] == "18KT WG"
    assert row["Order #"] == "ORDER123"
    assert row["Qty"] == 1
    assert row["Weight"] == 5.0
    assert row["ItemTypeID"] == 11
    assert row["C1:Item Name"] == "18KT WG"
    assert row["C1:Weight"] == 4.5
    assert row["C1: Total H.Cost"] == 200.0
    assert row["C2:Item Name"] == "Labor"
    assert row["C2: Total H.Cost"] == 30.0
    assert row["C3:Item Name"] == "Diamonds"
    assert row["C3:LotName"] == "CERT001"
    assert row["C3:Shape"] == "RD"
    assert row["C3: Stone Position"] == "C"


def test_build_rows_falls_back_to_master_style_jobsheet_lookup():
    item = _item(sn="STYLE1/2")
    jobsheet_index = {"STYLE1": {"Order Id": "ORDER123"}}

    rows, _ = fantasy_file.build_rows([item], jobsheet_index)

    assert rows[0]["Order #"] == "ORDER123"


def test_build_rows_warns_when_an_item_has_no_matching_jobsheet_row():
    rows, warnings = fantasy_file.build_rows([_item(sn="ORPHAN1")], {})

    assert rows[0]["Order #"] is None
    assert any("no matching jobsheet row" in w and "ORPHAN1" in w for w in warnings)


def test_build_rows_does_not_warn_when_every_item_matches_a_jobsheet_row():
    jobsheet_index = {"STYLE1": {"Order Id": "ORDER123"}}

    _, warnings = fantasy_file.build_rows([_item()], jobsheet_index)

    assert not any("jobsheet row" in w for w in warnings)


def test_build_rows_warns_when_a_kt_code_is_not_in_the_material_table():
    rows, warnings = fantasy_file.build_rows(
        [_item(kt="99KT-MADEUP", sn="ODDBALL1")], {"ODDBALL1": {"Order Id": "X"}}
    )

    assert rows[0]["Metal"] == "99KT-MADEUP"
    assert any(
        "99KT-MADEUP" in w and "ODDBALL1" in w and "FANTASY_MATERIAL_MAP" in w
        for w in warnings
    )


def test_build_rows_does_not_warn_when_every_kt_code_is_recognized():
    _, warnings = fantasy_file.build_rows([_item(kt="18KT WG")], {"STYLE1": {"Order Id": "X"}})

    assert not any("FANTASY_MATERIAL_MAP" in w for w in warnings)


def test_build_rows_lot_name_is_always_the_style_number():
    # Real-world regression: the source tool's LotName fallback chain
    # (settingCert || igiCert || sn) was found to disagree with a real
    # reference Open Stock file, which shows the style number on every row
    # checked even when a certificate was available -- LotName is always
    # sn regardless of item/stone cert data (user decision).
    rows, _ = fantasy_file.build_rows([_item(cert="ITEMCERT", stones=[_stone(cert="STONECERT")])], {})
    assert rows[0]["LotName"] == "STYLE1"

    rows2, _ = fantasy_file.build_rows([_item(sn="STYLE2", cert=None)], {})
    assert rows2[0]["LotName"] == "STYLE2"


def test_build_rows_excludes_side_stones_for_no_side_diamond_categories():
    item = _item(cat="BRACELET", stones=[
        _stone(position="center", cert="C1"),
        _stone(position="side", cert=None, shape="EM"),
    ])

    rows, _ = fantasy_file.build_rows([item], {})

    assert rows[0]["C3:Item Name"] == "Diamonds"
    assert rows[0]["C4:Item Name"] is None


def test_build_rows_warns_and_truncates_when_more_than_twenty_stone_groups():
    stones_list = [_stone(position="side", shape=f"SHAPE{i}", cert=None) for i in range(21)]
    item = _item(stones=stones_list)

    rows, warnings = fantasy_file.build_rows([item], {})

    assert any("truncated" in w for w in warnings)
    assert rows[0]["C22:Item Name"] is not None


from io import BytesIO

import openpyxl
from openpyxl.styles import PatternFill


def _rgb_of(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid").start_color.rgb


def test_write_xlsx_includes_the_header_row_and_row_values():
    rows = [{column: None for column in fantasy_file.FANTASY_COLUMNS}]
    rows[0]["Metal"] = "18KT WG"
    rows[0]["Qty"] = 2
    rows[0]["C3:Item Name"] = "Diamonds"

    workbook = openpyxl.load_workbook(BytesIO(fantasy_file.write_xlsx(rows)))
    sheet = workbook.active

    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header == fantasy_file.FANTASY_COLUMNS

    metal_col = fantasy_file.FANTASY_COLUMNS.index("Metal") + 1
    qty_col = fantasy_file.FANTASY_COLUMNS.index("Qty") + 1
    assert sheet.cell(row=2, column=metal_col).value == "18KT WG"
    assert sheet.cell(row=2, column=qty_col).value == 2


def test_write_xlsx_colors_match_a_real_reference_files_actual_formatting():
    # Colors verified directly against 4 real reference "Open Stock" files'
    # actual cell formatting (Data/JNE016, JNE008, JNE009, JNE017 Open
    # Stock.xls, local-only, real business data) -- all 4 agree exactly.
    rows = [{column: None for column in fantasy_file.FANTASY_COLUMNS}]

    workbook = openpyxl.load_workbook(BytesIO(fantasy_file.write_xlsx(rows)))
    sheet = workbook.active

    def fill_of(column_name):
        col_index = fantasy_file.FANTASY_COLUMNS.index(column_name) + 1
        return sheet.cell(row=1, column=col_index).fill.start_color.rgb

    assert fill_of("Item Name") == _rgb_of("C0C0C0")
    assert fill_of("ItemTypeID") == _rgb_of("C0C0C0")
    assert fill_of("C1:Item Name") == _rgb_of("FFFF99")
    assert fill_of("C2:Item Name") == _rgb_of("00CCFF")
    assert fill_of("C3:Item Name") == _rgb_of("FFCC99")
    assert fill_of("C4:Item Name") == _rgb_of("FFFF99")
    assert fill_of("C5:Item Name") == _rgb_of("FFFF00")
    assert fill_of("C6:Item Name") == _rgb_of("99CC00")
    assert fill_of("C7:Item Name") == _rgb_of("99CC00")

    for unfilled in ("PO Doc ID", "PO Doc Line", "C8:Item Name", "C22:Item Name"):
        col_index = fantasy_file.FANTASY_COLUMNS.index(unfilled) + 1
        assert sheet.cell(row=1, column=col_index).fill.fill_type is None
